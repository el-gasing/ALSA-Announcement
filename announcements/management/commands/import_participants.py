from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from announcements.departments import DEPARTMENT_NAME_TO_SLUG
from announcements.models import Participant


class Command(BaseCommand):
    help = 'Import participants from an Excel workbook into the database.'

    def add_arguments(self, parser):
        parser.add_argument(
            'workbook',
            nargs='?',
            default='data.xlsx',
            help='Path to the Excel workbook to import.',
        )

    def handle(self, *args, **options):
        workbook_path = Path(options['workbook']).expanduser().resolve()
        if not workbook_path.exists():
            raise CommandError(f'Workbook not found: {workbook_path}')

        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError('Workbook is empty.')

        headers = [str(cell).strip().lower() if cell is not None else '' for cell in rows[0]]
        expected = {'nim', 'nama', 'department'}
        if not expected.issubset(set(headers)):
            raise CommandError(
                f'Workbook must contain columns {sorted(expected)}. Found: {headers}'
            )

        idx = {header: headers.index(header) for header in expected}
        records = []
        seen_nims = set()

        for excel_row, row in enumerate(rows[1:], start=2):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            nim = str(row[idx['nim']] or '').strip().upper()
            name = str(row[idx['nama']] or '').strip()
            department_name = str(row[idx['department']] or '').strip()

            if not nim or not name or not department_name:
                raise CommandError(f'Row {excel_row} has empty required fields.')

            if nim in seen_nims:
                raise CommandError(f'Duplicate NIM in workbook: {nim} (row {excel_row})')
            seen_nims.add(nim)

            department_slug = DEPARTMENT_NAME_TO_SLUG.get(department_name)
            if not department_slug:
                raise CommandError(
                    f'Unknown department "{department_name}" on row {excel_row}.'
                )

            records.append(
                Participant(
                    nim=nim,
                    name=name,
                    department=department_slug,
                )
            )

        if not records:
            raise CommandError('No participant rows found to import.')

        with transaction.atomic():
            Participant.objects.all().delete()
            Participant.objects.bulk_create(records)

        self.stdout.write(
            self.style.SUCCESS(
                f'Imported {len(records)} participants from {workbook_path.name}.'
            )
        )
